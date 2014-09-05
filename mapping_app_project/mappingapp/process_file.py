# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from mappingapp.extractosl import get_osl_sample_info
from mappingapp.extractC14 import get_C14_sample_info
from mappingapp.extracttcn import get_tcn_sample_info
from mappingapp.extractsite import get_site_info, columns
from mappingapp.conversion import convert_date, convert_lat_long
import datetime


# does the sheet have a sample code?
def get_sample_code(sheet):
    for row in sheet.iter_rows():
        id = None
        for cell in row:
            if cell.value is not None and isinstance(cell.value, basestring):
                if 'Unique Sample Identifier' in cell.value:
                    col = columns.index(cell.column)
                    val_col = columns[col+1]
                    id = sheet[val_col + str(cell.row)].value
                    return id


# determine what type of sample, if any, is on a worksheet
def get_sheet_type(site_sheet):
    for row in site_sheet.iter_rows():
        for cell in row:
            if cell.value is not None and isinstance(cell.value, basestring):
                if 'Section A: Site Information' in cell.value:
                    sheet = 'osl_site'
                    return sheet
                elif 'Site Information' in cell.value and 'Section A' not in cell.value:
                    sheet = 'standard_site'
                    return sheet
                elif 'TCN' in cell.value:
                    sheet = 'TCN_Sample'
                    return sheet
                elif 'OSL' in cell.value:
                    sheet = 'OSL_Sample'
                    return sheet
                elif '14C' in cell.value:
                    sheet = '14C_Sample'
                    return sheet
    return None


# process a complete file
def process_file(filename):

    # dictionary for the data from all samples
    samples = {}

    # counter to track the number of samples
    counter = 0

    # create workbook from file using Openpyxl
    wb = load_workbook(filename, use_iterators=True)

    # get list of the names of the sheets
    sheet_names = wb.get_sheet_names()

    # set site name to None
    site_name = None

    # determine type of sheet and process accordingly.  For sample sheets check that a sample code has been entered
    # Ignore those without sample codes.
    #  Add returning dictionary elements to combined samples dict and increment counter if processed
    for sheet in sheet_names:
        site_sheet = wb[sheet]
        type = get_sheet_type(site_sheet)
        id = get_sample_code(site_sheet)

        if type == 'standard_site':
            site_name = get_site_info(site_sheet, 'standard')

        elif type == 'osl_site':
            site_name = get_site_info(site_sheet, 'osl')

        elif type == '14C_Sample':
            if id is not None:
                counter += 1
                results = get_C14_sample_info(site_sheet, counter)
                for k, v in results.iteritems():
                    samples[k] = v

        elif type == 'OSL_Sample':
            if id is not None:
                counter += 1
                results = get_osl_sample_info(site_sheet, counter)
                for k, v in results.iteritems():
                    samples[k] = v

        elif type == 'TCN_Sample':
            if id is not None:
                counter += 1
                results = get_tcn_sample_info(site_sheet, counter)
                for k, v in results.iteritems():
                    samples[k] = v

    # add the total sample count and site name to the dictionary for return
    samples['sample_count'] = counter
    samples['site_name'] = site_name

    if counter == 0:
        return None
    else:
        return samples

