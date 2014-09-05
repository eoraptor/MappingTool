# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from mappingapp.models import Sample_Site, Coordinates
from mappingapp.extractosl import get_osl_sample_info
from mappingapp.extractC14 import get_C14_sample_info
from mappingapp.extracttcn import get_tcn_sample_info
from mappingapp.conversion import convert_date, convert_lat_long
import datetime

# column headings
columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
           'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']


# iterate over Standard site sheet to determine which cells to extract data from
def get_site_cell_positions(ws):

    # dictionary of field name cells
    positions = {'Site Name: ': '', 'Location (inc. Transect no.)': '', 'Latitude(decimal deg)': '',
                 'Longtitude(decimal deg)(West is -ve)': '', 'BNG/ING?': '', 'Northing:': '', 'Easting:': '',
                 'Elevation             (m ASL)': '', 'Date Sampled:': '', 'Geomorph Setting:': '',
                 'Type of Samples collected (TCN/14C/OSL)': '', 'Collected by:': '', 'Photographs Taken (Y/N):': '',
                 'Photo labels/Time stamps:': '', 'Notes': ''}

    # iterate over sheet to find positions of value cells
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            # remove newlines
            if val is not None and isinstance(val, basestring):
                val = val.replace('\n', '')

                if val in positions:
                    # all value cells one to the right of their field name cells
                    col = columns.index(cell.column)
                    val_col = columns[col+1]
                    positions[val] = val_col + str(cell.row)

    return positions


# get OSL site cell positions - Sheet in different format from other site sheets
def get_osl_site_cell_positions(ws):

    # OSL site field names
    osl_positions = {'Site Name: ': '', 'Location (inc. Transect no.)': '', 'Latitude': '',
                     'Longtitude (East +ve)': '', 'BNG (Easting)': '', 'Northing': '', 'ING (Easting)': '',
                     'Elevation (m asl)': '', 'Date Sampled': '', 'Geomorph Setting': '',
                     'Type of Samples collected (TCN/14C/OSL)': '', 'Collected by': '', 'Photographs Taken (Y/N)': '',
                     'Photo labels/Time stamps': '', 'Notes': '', 'BNG/ING?': ''}

    # iterate over sheet to find value cell locations
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            # remove newlines
            if val is not None and isinstance(val, basestring):
                val = val.replace('\n', '')

                if val in osl_positions:
                    # notes value in cell below
                    if val == 'Notes':
                        osl_positions[val] = cell.column + str(cell.row+1)

                    # site name in cell to the right but row above due to cell splitting
                    elif val == 'Site Name: ':
                        col = columns.index(cell.column)
                        val_col = columns[col+1]
                        osl_positions[val] = val_col + str(cell.row-1)

                    # all others one to the right
                    else:
                        col = columns.index(cell.column)
                        val_col = columns[col+1]
                        osl_positions[val] = val_col + str(cell.row)

    # BNG or ING data must be calculated on the basis of which of the Easting cells were completed.
    northing_cell = ''
    easting_cell = ''

    # BNG --> Northing cell is three columns across from the BNG easting cell, BNG/ING value = BNG
    if ws[osl_positions['BNG (Easting)']].value is not None:
        easting_cell = osl_positions['BNG (Easting)']
        northing_cell = columns[columns.index(easting_cell[0])+3] + easting_cell[1]
        osl_positions['BNG/ING?'] = 'BNG'

    # ING --> Northing cell is three columns across from the ING easting cell, BNG/ING value = ING
    elif ws[osl_positions['ING (Easting)']].value is not None:
        easting_cell = osl_positions['ING (Easting)']
        northing_cell = columns[columns.index(easting_cell[0])+3] + easting_cell[1]
        osl_positions['BNG/ING?'] = 'ING'

    # return dictionary of cell positions in same format as that of Standard site sheets but including BNG/ING value
    # as opposed to cell location
    positions = {'Site Name: ': osl_positions['Site Name: '],
                 'Location (inc. Transect no.)': osl_positions['Location (inc. Transect no.)'],
                 'Latitude(decimal deg)': osl_positions['Latitude'],
                 'Longtitude(decimal deg)(West is -ve)': osl_positions['Longtitude (East +ve)'],
                 'BNG/ING?': osl_positions['BNG/ING?'],
                 'Northing:': northing_cell, 'Easting:': easting_cell,
                 'Elevation             (m ASL)': osl_positions['Elevation (m asl)'],
                 'Date Sampled:': osl_positions['Date Sampled'],
                 'Geomorph Setting:': osl_positions['Geomorph Setting'],
                 'Type of Samples collected (TCN/14C/OSL)': osl_positions['Type of Samples collected (TCN/14C/OSL)'],
                 'Collected by:': osl_positions['Collected by'],
                 'Photographs Taken (Y/N):': osl_positions['Photographs Taken (Y/N)'],
                 'Photo labels/Time stamps:': osl_positions['Photo labels/Time stamps'],
                 'Notes': osl_positions['Notes']}

    return positions


# extract the data from the site sheet
def get_site_info(site_sheet, type):

    # get the positions of the value cells
    if type == 'osl':
        positions = get_osl_site_cell_positions(site_sheet)
        bng_ing = positions['BNG/ING?']
    else:
        positions = get_site_cell_positions(site_sheet)
        bng_ing = site_sheet[positions['BNG/ING?']].value

    # create variables
    site_name = site_sheet[positions['Site Name: ']].value
    site_location = site_sheet[positions['Location (inc. Transect no.)']].value
    site_latitude = site_sheet[positions['Latitude(decimal deg)']].value
    site_longitude = site_sheet[positions['Longtitude(decimal deg)(West is -ve)']].value
    site_northing = site_sheet[positions['Northing:']].value
    site_easting = site_sheet[positions['Easting:']].value
    site_elevation = site_sheet[positions['Elevation             (m ASL)']].value
    site_date = site_sheet[positions['Date Sampled:']].value
    geomorph = site_sheet[positions['Geomorph Setting:']].value
    type = site_sheet[positions['Type of Samples collected (TCN/14C/OSL)']].value
    collector = site_sheet[positions['Collected by:']].value
    photographs = site_sheet[positions['Photographs Taken (Y/N):']].value
    photo_labels = site_sheet[positions['Photo labels/Time stamps:']].value
    site_notes = site_sheet[positions['Notes']].value

    # remove newlines from notes
    if site_notes is not None:
        site_notes = site_notes.replace('\n', ' ')
        site_notes = ' '.join(site_notes.split())

    # convert photographs into boolean.  If not yes/no, add to site notes
    if photographs is not None:
        if len(photographs) > 3:
            if site_notes is not None:
                site_notes = site_notes + photographs + '. '
                photographs = None
            else:
                site_notes = photographs + '. '
                photographs = None
        else:
            photographs = photographs.lower()
            if photographs.startswith('y'):
                photographs = True
            elif photographs.startswith('n'):
                photographs = False


    # convert date if format incorrect
    if site_date is not None:

        # convert to string
        date = str(site_date)

        # remove time
        if ' 00:00:00' in date:
            date = date.replace(' 00:00:00', '')

        try:
            # format accepted by database
            site_date = datetime.datetime.strptime(date, '%Y-%m-%d')

        except:
            # try conversion - if fails, add to notes and set date to None
            if '.' in date:
                try:
                    site_date = convert_date(date)
                    if site_date == 'Error' and site_notes is not None:
                        site_notes = site_notes + ' ' + date + '. '
                        site_date = None
                    elif site_date == 'Error' and site_notes is None:
                        site_notes = date + '. '
                        site_date = None
                    else:
                        date = datetime.datetime.strptime(site_date, '%d/%m/%Y')
                        site_date = str(date.strftime('%Y-%m-%d'))
                except:
                    if site_notes is not None:
                        site_notes = site_notes + ' ' + date + '. '
                    else:
                        site_notes = date + '. '
                    site_date = None
            else:
                # unrecognisable format - add to notes and set date to None
                if site_notes is not None:
                    site_notes = site_notes + ' ' + date + '. '
                else:
                    site_notes = date + '. '
                site_date = None

    # convert northing and easting to integers, set to None if non-numerical
    if site_northing is not None:
        try:
            site_northing = int(site_northing)
        except:
            site_northing = None

    if site_easting is not None:
        try:
            site_easting = int(site_easting)
        except:
            site_easting = None

    # convert Lat/Long if not floats, set to None if fails
    if site_latitude is not None and not isinstance(site_latitude, (float, int)):
            site_latitude = convert_lat_long(site_latitude)
            if not isinstance(site_latitude, float):
                site_latitude = None

    if site_longitude is not None and not isinstance(site_longitude, (float, int)):
            site_longitude = convert_lat_long(site_longitude)
            if isinstance(site_longitude, float):
                site_longitude = -1 * site_longitude
            else:
                site_longitude = None



    site_coordinates = None
    site = None

    # all fields empty - do nothing
    if site_name is None and site_location is None and site_date is None and geomorph is None and type is None and\
                    photographs is None and photo_labels is None and site_notes is None and\
                    site_coordinates is None and collector is None:
        pass

    else:
        # does site already exist?
        try:
            site = Sample_Site.objects.get(site_name=site_name)
        except:
            pass

        # doesn't exist --> create new site
        if site is None:

            # create coordinates instance if values
            if bng_ing is None and site_easting is None and site_northing is None and site_latitude is None and\
                    site_longitude is None and site_elevation is None:
                pass

            else:
                site_coordinates = Coordinates.objects.create(bng_ing=bng_ing, grid_reference=None,
                                                              easting=site_easting, northing=site_northing,
                                                              latitude=site_latitude, longitude=site_longitude,
                                                              elevation=site_elevation)

            # create site
            Sample_Site.objects.create(site_name=site_name, site_location=site_location, county=None,
                                                 site_date=site_date, operator=None, geomorph_setting=geomorph,
                                                 sample_type_collected=type, photos_taken=photographs,
                                                 photographs=photo_labels, site_notes=site_notes,
                                                 site_coordinates=site_coordinates, collected_by=collector)
    # return name of site only
    return site_name






