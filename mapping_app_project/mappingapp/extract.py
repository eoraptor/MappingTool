# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from mappingapp.models import Sample_Site, Coordinates
from mappingapp.extractosl import get_osl_sample_info
from mappingapp.extractC14 import get_C14_sample_info
from mappingapp.extracttcn import get_tcn_sample_info
from mappingapp.conversion import convert_date, convert_lat_long
import datetime


columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
            'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']


# iterate over the site sheet to determine which cells to extract data from
def get_site_cell_positions(ws):

    positions = {'Site Name: ': '', 'Location (inc. Transect no.)': '', 'Latitude(decimal deg)': '',
                 'Longtitude(decimal deg)(West is -ve)': '', 'BNG/ING?': '', 'Northing:': '', 'Easting:': '',
                 'Elevation             (m ASL)': '', 'Date Sampled:': '', 'Geomorph Setting:': '',
                 'Type of Samples collected (TCN/14C/OSL)': '', 'Collected by:': '', 'Photographs Taken (Y/N):': '',
                 'Photo labels/Time stamps:': '', 'Notes': ''}

    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if val is not None and isinstance(val, basestring):
                val = val.replace('\n', '')

                if val in positions:
                    col = columns.index(cell.column)
                    val_col = columns[col+1]
                    positions[val] = val_col + str(cell.row)

    return positions


# get OSL sample site cell positions
def get_osl_site_cell_positions(ws):

    osl_positions = {'Site Name: ': '', 'Location (inc. Transect no.)': '', 'Latitude': '',
             'Longtitude (East +ve)': '', 'BNG (Easting)': '', 'Northing': '', 'ING (Easting)': '',
             'Elevation (m asl)': '', 'Date Sampled': '', 'Geomorph Setting': '',
             'Type of Samples collected (TCN/14C/OSL)': '', 'Collected by': '', 'Photographs Taken (Y/N)': '',
             'Photo labels/Time stamps': '', 'Notes': '', 'BNG/ING?':''}

    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if val is not None and isinstance(val, basestring):
                val = val.replace('\n', '')

                if val in osl_positions:
                    if val == 'Notes':
                        osl_positions[val] = cell.column + str(cell.row+1)
                    elif val == 'Site Name: ':
                        col = columns.index(cell.column)
                        val_col = columns[col+1]
                        osl_positions[val] = val_col + str(cell.row-1)
                    else:
                        col = columns.index(cell.column)
                        val_col = columns[col+1]
                        osl_positions[val] = val_col + str(cell.row)

    # BNG or ING data?
    if ws[osl_positions['BNG (Easting)']].value is not None:
        easting_cell = osl_positions['BNG (Easting)']
        northing_cell = columns[easting_cell.column+3] + str(easting_cell.row)
        osl_positions['BNG/ING?'] = 'BNG'
    elif ws[osl_positions['ING (Easting)']].value is not None:
        easting_cell = osl_positions['ING (Easting)']
        northing_cell = columns[columns.index(easting_cell[0])+3] + easting_cell[1]
        osl_positions['BNG/ING?'] = 'ING'

    positions = {'Site Name: ': osl_positions['Site Name: '],
                 'Location (inc. Transect no.)': osl_positions['Location (inc. Transect no.)'],
                 'Latitude(decimal deg)': osl_positions['Latitude'],
                 'Longtitude(decimal deg)(West is -ve)': osl_positions['Longtitude (East +ve)'],
                 'BNG/ING?': osl_positions['BNG/ING?'],
                 'Northing:': northing_cell, 'Easting:': easting_cell,
                 'Elevation             (m ASL)': osl_positions['Elevation (m asl)'],
                 'Date Sampled:': osl_positions['Date Sampled'],
                 'Geomorph Setting:': osl_positions['Geomorph Setting'],
                 'Type of Samples collected (TCN/14C/OSL)': osl_positions['Type of Samples collected (TCN/14C/OSL)'],
                 'Collected by:': osl_positions['Collected by'],
                 'Photographs Taken (Y/N):': osl_positions['Photographs Taken (Y/N)'],
                 'Photo labels/Time stamps:': osl_positions['Photo labels/Time stamps'],
                 'Notes': osl_positions['Notes']}

    return positions


# extract the data from the site sheet
def get_site_info(site_sheet, type):

    if type == 'osl':
        positions = get_osl_site_cell_positions(site_sheet)
        bng_ing = positions['BNG/ING?']
    else:
        positions = get_site_cell_positions(site_sheet)
        bng_ing = site_sheet[positions['BNG/ING?']].value

    site_name = site_sheet[positions['Site Name: ']].value
    site_location = site_sheet[positions['Location (inc. Transect no.)']].value
    site_latitude = site_sheet[positions['Latitude(decimal deg)']].value
    site_longitude = site_sheet[positions['Longtitude(decimal deg)(West is -ve)']].value
    site_northing = site_sheet[positions['Northing:']].value
    site_easting = site_sheet[positions['Easting:']].value
    site_elevation = site_sheet[positions['Elevation             (m ASL)']].value
    site_date = site_sheet[positions['Date Sampled:']].value
    geomorph = site_sheet[positions['Geomorph Setting:']].value
    type = site_sheet[positions['Type of Samples collected (TCN/14C/OSL)']].value
    collector = site_sheet[positions['Collected by:']].value
    photographs = site_sheet[positions['Photographs Taken (Y/N):']].value
    photo_labels = site_sheet[positions['Photo labels/Time stamps:']].value
    site_notes = site_sheet[positions['Notes']].value

    # remove newlines from notes
    if site_notes is not None:
        site_notes = site_notes.replace('\n', ' ')
        site_notes = ' '.join(site_notes.split())

    # convert photographs into boolean.  If not yes/no, add to site notes
    if photographs is not None:
        if len(photographs) > 3:
            if site_notes is not None:
                site_notes = site_notes + photographs + '. '
            else:
                site_notes = photographs + '. '
                photographs = None
        else:
            photographs = photographs.lower()
            if photographs.startswith('y'):
                photographs = True
            elif photographs.startswith('n'):
                photographs = False

    # convert date to correct format.  If notes add to site notes.
    if site_date is not None and not isinstance(site_date, datetime.date):
        if len(site_date) > 10:
            if site_notes is not None:
                site_notes = site_notes + site_date
                site_date = None
            else:
                site_notes = site_date
        else:
            date = str(site_date)
            if '.' in date:
                site_date = convert_date(date)

    if site_northing is not None:
        site_northing = int(site_northing)

    if site_easting is not None:
        site_easting = int(site_easting)

    if site_latitude is not None and not isinstance(site_latitude, float):
            site_latitude = convert_lat_long(site_latitude)

    if site_longitude is not None and not isinstance(site_longitude, float):
            site_longitude = convert_lat_long(site_longitude)
            if isinstance(site_longitude, float):
                site_longitude = -1 * site_longitude

    site_coordinates = None
    site = None

    if bng_ing is None and site_easting is None and site_northing is None and site_latitude is None and\
                    site_longitude is None and site_elevation is None:
        pass
    else:
        site_coordinates = Coordinates.objects.get_or_create(bng_ing=bng_ing, grid_reference=None, easting=site_easting,
                                                             northing=site_northing, latitude=site_latitude,
                                                             longitude=site_longitude, elevation=site_elevation)[0]

    if site_name is None and site_location is None and site_date is None and geomorph is None and type is None and\
                    photographs is None and photo_labels is None and site_notes is None and\
                    site_coordinates is None and collector is None:
        pass

    else:
        try:
            site = Sample_Site.objects.get(site_name=site_name)
        except:
            pass

        if site is None:
            Sample_Site.objects.create(site_name=site_name, site_location=site_location, county=None,
                                                 site_date=None, operator=None, geomorph_setting=geomorph,
                                                 sample_type_collected=type, photos_taken=photographs,
                                                 photographs=photo_labels, site_notes=site_notes,
                                                 site_coordinates=site_coordinates, collected_by=collector)

    return site_name


# does the sheet have a sample code?
def get_sample_code(sheet):
    for row in sheet.iter_rows():
        id = None
        for cell in row:
            if cell.value is not None and isinstance(cell.value, basestring):
                if 'Unique Sample Identifier' in cell.value:
                    print 'here'
                    col = columns.index(cell.column)
                    val_col = columns[col+1]
                    id = sheet[val_col + str(cell.row)].value
                    return id


# determine what type of sample, if any, is on a worksheet
def get_sheet_type(site_sheet):
    for row in site_sheet.iter_rows():
        for cell in row:
            if cell.value is not None and isinstance(cell.value, basestring):
                if 'Section A: Site Information' in cell.value:
                    sheet = 'osl_site'
                    return sheet
                elif 'Site Information' in cell.value and 'Section A' not in cell.value:
                    sheet = 'standard_site'
                    return sheet
                elif 'TCN' in cell.value:
                    sheet = 'TCN_Sample'
                    return sheet
                elif 'OSL' in cell.value:
                    sheet = 'OSL_Sample'
                    return sheet
                elif '14C' in cell.value:
                    sheet = '14C_Sample'
                    return sheet
    return None

# process a complete file
def process_file(filename):
    # need error handling

    site_name = None

    samples = {}

    counter = 0

    wb = load_workbook(filename, use_iterators=True)

    sheet_names = wb.get_sheet_names()
    site_name = None

    for sheet in sheet_names:
        site_sheet = wb[sheet]
        type = get_sheet_type(site_sheet)
        id = get_sample_code(site_sheet)

        if type == 'standard_site':
            site_name = get_site_info(site_sheet, 'standard')

        elif type == 'osl_site':
            site_name = get_site_info(site_sheet, 'osl')

        elif type == '14C_Sample':
            if id is not None:
                counter += 1
                results = get_C14_sample_info(site_sheet, counter)
                for k, v in results.iteritems():
                    samples[k] = v

        elif type == 'OSL_Sample':
            if id is not None:
                counter += 1
                results = get_osl_sample_info(site_sheet, counter)
                for k, v in results.iteritems():
                    samples[k] = v

        elif type == 'TCN_Sample':
            if id is not None:
                counter += 1
                results = get_tcn_sample_info(site_sheet, counter)
                for k, v in results.iteritems():
                    samples[k] = v

    samples['sample_count'] = counter
    samples['site_name'] = site_name

    if counter == 0:
        return None
    else:
        return samples




