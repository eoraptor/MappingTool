# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from conversion import convert_date, convert_lat_long
import datetime

columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N',
            'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']


def get_NERC_form_cell_positions(ws):

    positions = {'Unique sample identifier*': '', 'Location': '', 'Latitude        oN/S': '',
                 'Longitude oE/W': '', 'Nature of sample': '', 'Collection date': '', 'In situ environment': '',
                 'Stratigraphic position (cm)': '','Possible contamination': '', 'Sample collector': ''}

    # retrieve cells positions with values.  Cells are read-only so do not have .cell and .row attributes
    # used a loop instead to find the starting positions.
    for column in columns:
        i = 1
        while i <= 20:
            cell = column + str(i)
            try:
                val = ws[cell].value
                if val is not None and isinstance(val, basestring):
                    if val in positions:
                        # starting cell below
                        positions[val] = column + str(i+1)
            except:
                pass
            i += 1

    return positions


# process a single row
def process_row(sample_sheet, positions, sample_count):

    # list for errors
    errors = []

    # missing keys
    missing_keys  = []

    # create variables from spreadsheet values
    sample_code = sample_sheet[positions['Unique sample identifier*']].value
    if sample_code is None or sample_code == '':
        return None
    else:
        sample_date = sample_sheet[positions['Collection date']].value
        sample_location_name = sample_sheet[positions['Location']].value
        latitude = sample_sheet[positions['Latitude        oN/S']].value
        longitude = sample_sheet[positions['Longitude oE/W']].value
        position = sample_sheet[positions['Stratigraphic position (cm)']].value
        material = sample_sheet[positions['Nature of sample']].value
        setting = sample_sheet[positions['In situ environment']].value
        collector = sample_sheet[positions['Sample collector']].value
        contamination = sample_sheet[positions['Possible contamination']].value
        notes = None

       # convert date if format incorrect - add to notes if date cannot be resolved
        if sample_date is not None:
            date = str(sample_date)

            if ' 00:00:00' in date:
                 date = date.replace(' 00:00:00', '')

            if '.' in date:
                sample_date = convert_date(date)
                if sample_date == 'Error':
                    if notes is not None:
                        notes = notes + ' ' + sample_date + '. '
                    else:
                        notes = sample_date + '. '
                    sample_date = None
            else:
                try:
                    date = datetime.datetime.strptime(date, '%Y-%m-%d')
                    sample_date = date.strftime('%d/%m/%Y')

                except:

                    if notes is not None:
                        notes = notes + ' ' + date + '. '
                    else:
                        notes = date + '. '
                    sample_date = None

        # convert latitude and longitude if format incorrect
        if latitude is not None and not isinstance(latitude, float):
            latitude = convert_lat_long(latitude)
            if latitude == 'Error':
                errors.append('Latitude')
                latitude = None

        if longitude is not None and not isinstance(longitude, float):
            longitude = convert_lat_long(longitude)
            if longitude == 'Error':
                errors.append('Longitude')
                longitude = None
            else:
                longitude = -1 * longitude

        sample_counter = str(sample_count)

        sample_details = {'sample_latitude'+sample_counter: latitude,
                          'sample_longitude'+sample_counter: longitude,
                          'sample_code'+sample_counter: sample_code,
                          'sample_location_name'+sample_counter: sample_location_name,
                          'sample_date'+sample_counter: sample_date,
                          'collector'+sample_counter: collector,
                          'material'+sample_counter: material,
                          'setting'+sample_counter: setting,
                          'contamination'+sample_counter: contamination,
                          'sample_notes'+sample_counter: notes, 'missing_keys'+sample_counter: missing_keys,
                          'position'+sample_counter: position, 'errors'+sample_counter: errors}

        return sample_details


def check_sheet_type(site_sheet):
    for row in site_sheet.iter_rows():
        for cell in row:
            if cell.value is not None and isinstance(cell.value, basestring):
                if 'SUBMITTER INFORMATION' in cell.value:
                    return True
    return False



# extract the data from a tcn sample sheet
def get_NERC_form_info(sample_sheet, sample_counter):

    positions = get_NERC_form_cell_positions(sample_sheet)

    results = {}

    MAX_ROWS = 14
    rows_checked = 1

    # get row values as long as a sample code is present
    while rows_checked <= MAX_ROWS:
        values = process_row(sample_sheet, positions, sample_counter)
        if values is not None:
            results.update(values)
            sample_counter += 1
        rows_checked += 1
        positions = increment_positions(positions)


    if results != {}:
        results['sample_count'] = sample_counter
        return results
    else:
        return None


# set cell targets to next row
def increment_positions(positions):
    for key in positions:
        if key != 'Possible contamination':
            cell = positions[key]
            next_row = int(cell[1:])+1
            positions[key] = cell[0] + str(next_row)

    return positions


def process_nerc_file(filename):

    wb = load_workbook(filename, use_iterators=True)
    sheet_names = wb.get_sheet_names()

    results = {}

    sample_counter = 1

    for name in sheet_names:
        sheet = wb[name]
        correct_type = check_sheet_type(sheet)
        if correct_type is True:
            data = get_NERC_form_info(sheet, sample_counter)
            if data is not None:
                results.update(data)
                sample_counter = results['sample_count']

    if sample_counter != 1:
        results['sample_count'] -= 1
        return results
    else:
        return None


